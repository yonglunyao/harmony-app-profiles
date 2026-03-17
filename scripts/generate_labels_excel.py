"""
生成HarmonyOS应用画像标签Excel文件
包含标签列表和检测方案矩阵
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 标签数据定义
LABELS_DATA = {
    "应用基本信息画像": {
        "APP_001": {"name": "bundle_name格式异常", "desc": "bundle_name不符合命名规范", "risk": "low", "weight": 1},
        "APP_002": {"name": "版本号格式异常", "desc": "versionCode或versionName格式异常", "risk": "low", "weight": 1},
        "APP_003": {"name": "签名信息缺失", "desc": "应用缺少签名或签名无效", "risk": "critical", "weight": 10},
        "APP_004": {"name": "证书链异常", "desc": "证书链不完整或过期", "risk": "high", "weight": 5},
        "APP_005": {"name": "应用图标异常", "desc": "图标缺失或包含违规内容", "risk": "medium", "weight": 3},
        "APP_006": {"name": "应用名称含敏感词", "desc": "应用名称包含敏感/违规词汇", "risk": "high", "weight": 4},
        "APP_007": {"name": "应用描述含诱导信息", "desc": "应用描述包含诱导下载/付费内容", "risk": "medium", "weight": 3},
        "APP_008": {"name": "权限声明与功能不符", "desc": "module.json5权限声明与实际功能不匹配", "risk": "medium", "weight": 3},
        "APP_009": {"name": "目标API版本过低", "desc": "targetAPI版本低于推荐值", "risk": "low", "weight": 2},
        "APP_010": {"name": "minAPI版本异常", "desc": "minAPIVersion设置异常", "risk": "low", "weight": 1},
        "APP_011": {"name": "安装包体积异常", "desc": "HAP包体积明显超过同类应用均值", "risk": "medium", "weight": 2},
        "APP_012": {"name": "安装包包含大量资源", "desc": "resources目录资源文件数量异常", "risk": "low", "weight": 1},
        "APP_013": {"name": "包含native库", "desc": "应用包含.so动态库", "risk": "medium", "weight": 2},
        "APP_014": {"name": "多HAP包结构", "desc": "应用采用多HAP模块化架构", "risk": "low", "weight": 1},
        "APP_015": {"name": "包含HarmonySharedLibrary", "desc": "应用包含共享库模块", "risk": "low", "weight": 1},
        "CS_001": {"name": "代码混淆不足", "desc": "代码混淆度低，易被逆向", "risk": "medium", "weight": 3},
        "CS_002": {"name": "包含调试信息", "desc": "发布版本包含调试符号", "risk": "medium", "weight": 3},
        "CS_003": {"name": "代码完整性异常", "desc": "代码存在被篡改痕迹", "risk": "critical", "weight": 9},
        "CS_004": {"name": "重打包检测", "desc": "应用为重新打包版本", "risk": "high", "weight": 7},
        "CS_005": {"name": "包含恶意Shell代码", "desc": "包含恶意Shell脚本", "risk": "critical", "weight": 10},
        "CS_006": {"name": "疑似木马代码", "desc": "包含木马特征代码", "risk": "critical", "weight": 10},
        "CS_007": {"name": "疑似蠕虫代码", "desc": "包含蠕虫传播代码", "risk": "critical", "weight": 10},
        "CS_008": {"name": "疑似间谍软件代码", "desc": "包含间谍软件特征", "risk": "critical", "weight": 10},
        "CS_009": {"name": "疑似勒索软件代码", "desc": "包含勒索软件特征", "risk": "critical", "weight": 10},
        "CS_010": {"name": "疑似挖矿代码", "desc": "包含加密货币挖矿代码", "risk": "high", "weight": 8},
        "CS_011": {"name": "使用已知漏洞SDK", "desc": "使用存在已知漏洞的SDK", "risk": "high", "weight": 7},
        "CS_012": {"name": "使用恶意第三方库", "desc": "使用被标记为恶意的第三方库", "risk": "critical", "weight": 9},
        "CS_013": {"name": "开源组件版本过低", "desc": "使用过时的开源组件版本", "risk": "medium", "weight": 4},
        "CS_014": {"name": "未声明第三方SDK", "desc": "使用但未声明第三方SDK", "risk": "medium", "weight": 4},
        "CS_015": {"name": "SDK异常初始化", "desc": "SDK初始化方式存在风险", "risk": "medium", "weight": 4},
    },
    "应用行为画像": {
        "BH_001": {"name": "恶意订阅检测", "desc": "应用存在自动订阅付费服务行为", "risk": "critical", "weight": 10},
        "BH_002": {"name": "隐私扣费", "desc": "在用户不知情下产生扣费行为", "risk": "critical", "weight": 10},
        "BH_003": {"name": "话费消耗异常", "desc": "异常消耗用户话费/流量", "risk": "high", "weight": 8},
        "BH_004": {"name": "支付密码窃取", "desc": "尝试窃取用户支付密码", "risk": "critical", "weight": 10},
        "BH_005": {"name": "订阅难以取消", "desc": "订阅服务取消流程异常复杂", "risk": "high", "weight": 6},
        "BH_006": {"name": "通讯录窃取", "desc": "非授权读取用户通讯录", "risk": "critical", "weight": 9},
        "BH_007": {"name": "短信记录窃取", "desc": "非授权读取用户短信记录", "risk": "critical", "weight": 9},
        "BH_008": {"name": "位置信息窃取", "desc": "后台频繁获取用户位置", "risk": "high", "weight": 7},
        "BH_009": {"name": "相册/文件窃取", "desc": "非授权访问用户相册或文件", "risk": "high", "weight": 7},
        "BH_010": {"name": "剪贴板窃取", "desc": "频繁读取剪贴板内容", "risk": "medium", "weight": 5},
        "BH_011": {"name": "后台偷跑流量", "desc": "应用在后台异常消耗流量", "risk": "high", "weight": 6},
        "BH_012": {"name": "恶意唤醒", "desc": "应用被其他应用频繁异常唤醒", "risk": "medium", "weight": 5},
        "BH_013": {"name": "持久化后台运行", "desc": "应用通过多种方式保持后台运行", "risk": "medium", "weight": 4},
        "BH_014": {"name": "CPU资源滥用", "desc": "应用异常占用CPU资源", "risk": "medium", "weight": 5},
        "BH_015": {"name": "内存泄漏", "desc": "应用存在明显内存泄漏问题", "risk": "low", "weight": 2},
        "BH_016": {"name": "强制推送广告", "desc": "强制用户观看广告才能使用功能", "risk": "high", "weight": 6},
        "BH_017": {"name": "恶意广告点击", "desc": "应用模拟用户点击广告", "risk": "critical", "weight": 8},
        "BH_018": {"name": "广告遮挡内容", "desc": "广告遮挡主要功能区域", "risk": "medium", "weight": 4},
        "BH_019": {"name": "广告欺诈", "desc": "展示虚假广告或欺骗性广告", "risk": "high", "weight": 6},
        "BH_020": {"name": "第三方广告SDK异常", "desc": "接入异常的第三方广告SDK", "risk": "high", "weight": 5},
        "BH_021": {"name": "静默安装应用", "desc": "应用静默安装其他应用", "risk": "critical", "weight": 10},
        "BH_022": {"name": "恶意快捷方式", "desc": "未经用户同意创建桌面快捷方式", "risk": "medium", "weight": 4},
        "BH_023": {"name": "恶意推送通知", "desc": "发送恶意或欺诈性通知", "risk": "high", "weight": 5},
        "BH_024": {"name": "应用推广欺诈", "desc": "通过诱导方式推广其他应用", "risk": "medium", "weight": 4},
        "BH_025": {"name": "色情/赌博内容推广", "desc": "推广色情、赌博等违法违规内容", "risk": "critical", "weight": 10},
        "NET_001": {"name": "明文HTTP通信", "desc": "使用HTTP而非HTTPS传输敏感数据", "risk": "high", "weight": 6},
        "NET_002": {"name": "SSL证书验证异常", "desc": "SSL证书验证被禁用或异常处理", "risk": "high", "weight": 7},
        "NET_003": {"name": "弱加密算法", "desc": "使用已弃用的弱加密算法", "risk": "medium", "weight": 4},
        "NET_004": {"name": "自签名证书", "desc": "使用自签名证书进行通信", "risk": "medium", "weight": 3},
        "NET_005": {"name": "证书固定缺失", "desc": "高风险通信未使用证书固定", "risk": "medium", "weight": 3},
        "NET_006": {"name": "连接恶意域名", "desc": "连接已知恶意域名或IP", "risk": "critical", "weight": 9},
        "NET_007": {"name": "连接境外可疑服务器", "desc": "频繁连接境外高风险服务器", "risk": "medium", "weight": 5},
        "NET_008": {"name": "异常网络请求模式", "desc": "请求模式呈现自动化特征", "risk": "high", "weight": 6},
        "NET_009": {"name": "隐私数据明文传输", "desc": "用户隐私数据以明文形式传输", "risk": "critical", "weight": 8},
        "NET_010": {"name": "DGA域名检测", "desc": "使用域名生成算法生成域名", "risk": "high", "weight": 7},
        "NET_011": {"name": "设备指纹采集", "desc": "采集设备唯一标识符", "risk": "medium", "weight": 4},
        "NET_012": {"name": "用户行为追踪", "desc": "长期追踪用户行为数据", "risk": "medium", "weight": 4},
        "NET_013": {"name": "数据跨境传输", "desc": "将用户数据传输至境外", "risk": "high", "weight": 5},
        "NET_014": {"name": "与三方共享数据", "desc": "未经用户同意与第三方共享数据", "risk": "high", "weight": 6},
        "NET_015": {"name": "数据压缩异常", "desc": "使用异常压缩方式隐藏数据", "risk": "medium", "weight": 4},
        "DS_001": {"name": "密码明文存储", "desc": "用户密码以明文形式存储", "risk": "critical", "weight": 10},
        "DS_002": {"name": "敏感数据明文存储", "desc": "敏感数据未加密存储", "risk": "high", "weight": 7},
        "DS_003": {"name": "数据存储位置异常", "desc": "数据存储在不安全位置", "risk": "medium", "weight": 4},
        "DS_004": {"name": "数据备份泄露风险", "desc": "应用数据备份包含敏感信息", "risk": "medium", "weight": 5},
        "DS_005": {"name": "删除不彻底", "desc": "用户删除应用后数据残留", "risk": "medium", "weight": 4},
        "DS_006": {"name": "使用硬编码密钥", "desc": "密钥硬编码在代码中", "risk": "critical", "weight": 8},
        "DS_007": {"name": "弱加密算法", "desc": "使用DES/RC4等弱加密算法", "risk": "high", "weight": 6},
        "DS_008": {"name": "加密模式不安全", "desc": "使用ECB等不安全加密模式", "risk": "high", "weight": 6},
        "DS_009": {"name": "IV/Nonce重用", "desc": "加密时重用初始化向量", "risk": "high", "weight": 5},
        "DS_010": {"name": "无数据加密", "desc": "敏感数据未进行任何加密处理", "risk": "high", "weight": 7},
        "DS_011": {"name": "越权访问数据", "desc": "应用可访问其他应用数据", "risk": "high", "weight": 6},
        "DS_012": {"name": "SQL注入风险", "desc": "存在SQL注入漏洞", "risk": "critical", "weight": 9},
        "DS_013": {"name": "NoSQL注入风险", "desc": "存在NoSQL注入漏洞", "risk": "high", "weight": 7},
        "DS_014": {"name": "路径遍历漏洞", "desc": "存在路径遍历攻击风险", "risk": "high", "weight": 7},
        "DS_015": {"name": "数据未脱敏展示", "desc": "敏感数据在界面未脱敏展示", "risk": "medium", "weight": 4},
        "PERM_001": {"name": "非必要权限申请", "desc": "申请与功能无关的敏感权限", "risk": "high", "weight": 6},
        "PERM_002": {"name": "过度权限申请", "desc": "申请权限数量明显超过同类应用", "risk": "medium", "weight": 4},
        "PERM_003": {"name": "敏感权限申请未说明", "desc": "申请敏感权限时未说明用途", "risk": "medium", "weight": 4},
        "PERM_004": {"name": "权限申请时机异常", "desc": "在非必要时机申请敏感权限", "risk": "medium", "weight": 3},
        "PERM_005": {"name": "系统签名权限申请", "desc": "申请system_grade权限", "risk": "high", "weight": 7},
        "PERM_006": {"name": "权限滥用", "desc": "申请权限但用于非声明目的", "risk": "high", "weight": 7},
        "PERM_007": {"name": "后台使用敏感权限", "desc": "在后台使用需要前台权限", "risk": "medium", "weight": 5},
        "PERM_008": {"name": "权限过度收集", "desc": "使用权限收集过度数据", "risk": "high", "weight": 6},
        "PERM_009": {"name": "持续使用位置权限", "desc": "持续后台获取位置信息", "risk": "high", "weight": 6},
        "PERM_010": {"name": "权限共享", "desc": "将权限获取的数据与第三方共享", "risk": "high", "weight": 6},
        "PERM_011": {"name": "设备管理器权限滥用", "desc": "滥用设备管理器权限", "risk": "critical", "weight": 8},
        "PERM_012": {"name": "无障碍服务滥用", "desc": "滥用无障碍服务权限", "risk": "critical", "weight": 8},
        "PERM_013": {"name": "悬浮窗滥用", "desc": "滥用悬浮窗权限进行恶意操作", "risk": "high", "weight": 6},
        "PERM_014": {"name": "通知监听滥用", "desc": "滥用通知监听权限", "risk": "high", "weight": 6},
        "PERM_015": {"name": "录音/相机滥用", "desc": "后台滥用录音或相机权限", "risk": "critical", "weight": 9},
        "COMP_001": {"name": "涉黄内容", "desc": "应用包含色情内容", "risk": "critical", "weight": 10},
        "COMP_002": {"name": "涉赌内容", "desc": "应用包含赌博内容", "risk": "critical", "weight": 10},
        "COMP_003": {"name": "涉政违规内容", "desc": "应用包含违法违规政治内容", "risk": "critical", "weight": 10},
        "COMP_004": {"name": "暴力恐怖内容", "desc": "应用包含暴力恐怖内容", "risk": "critical", "weight": 10},
        "COMP_005": {"name": "诈骗内容", "desc": "应用包含诈骗信息", "risk": "critical", "weight": 9},
        "COMP_006": {"name": "缺少隐私政策", "desc": "应用缺少隐私政策声明", "risk": "high", "weight": 6},
        "COMP_007": {"name": "隐私政策不合规", "desc": "隐私政策内容不完整或存在违规", "risk": "high", "weight": 7},
        "COMP_008": {"name": "未符合儿童隐私保护", "desc": "未符合儿童隐私保护规定", "risk": "high", "weight": 7},
        "COMP_009": {"name": "未符合数据出境规定", "desc": "数据出境未符合监管要求", "risk": "critical", "weight": 8},
        "COMP_010": {"name": "违反行业监管要求", "desc": "违反特定行业监管要求", "risk": "high", "weight": 7},
    },
    "应用舆情画像": {
        "PB_001": {"name": "恶意举报集中", "desc": "收到大量用户恶意举报", "risk": "high", "weight": 6},
        "PB_002": {"name": "负面评价占比高", "desc": "用户负面评价比例异常高", "risk": "medium", "weight": 4},
        "PB_003": {"name": "投诉扣费问题", "desc": "多个用户投诉扣费问题", "risk": "high", "weight": 7},
        "PB_004": {"name": "投诉隐私问题", "desc": "多个用户投诉隐私泄露", "risk": "high", "weight": 7},
        "PB_005": {"name": "疑似刷好评", "desc": "存在刷好评行为迹象", "risk": "medium", "weight": 3},
        "PB_006": {"name": "涉及负面新闻", "desc": "应用或开发者涉及负面新闻", "risk": "high", "weight": 6},
        "PB_007": {"name": "安全事件关联", "desc": "应用关联到已知安全事件", "risk": "critical", "weight": 8},
        "PB_008": {"name": "监管机构通报", "desc": "被监管机构通报或处罚", "risk": "critical", "weight": 9},
        "PB_009": {"name": "媒体负面报道", "desc": "被权威媒体负面报道", "risk": "high", "weight": 6},
        "PB_010": {"name": "社交媒体负面舆情", "desc": "社交媒体出现大规模负面讨论", "risk": "medium", "weight": 5},
        "REP_001": {"name": "病毒报毒率高", "desc": "被多款杀毒软件报毒", "risk": "critical", "weight": 9},
        "REP_002": {"name": "威胁情报标记", "desc": "被威胁情报平台标记为恶意", "risk": "critical", "weight": 10},
        "REP_003": {"name": "恶意样本库匹配", "desc": "匹配到已知恶意样本特征", "risk": "critical", "weight": 10},
        "REP_004": {"name": "灰度样本标记", "desc": "被标记为灰度或可疑样本", "risk": "high", "weight": 7},
        "REP_005": {"name": "样本家族关联", "desc": "关联到已知恶意家族", "risk": "high", "weight": 8},
        "REP_006": {"name": "IP信誉低", "desc": "通信IP信誉分数低", "risk": "high", "weight": 6},
        "REP_007": {"name": "域名信誉低", "desc": "通信域名信誉分数低", "risk": "high", "weight": 6},
        "REP_008": {"name": "关联恶意网络", "desc": "关联到已知恶意网络", "risk": "critical", "weight": 8},
        "REP_009": {"name": "CDN滥用", "desc": "滥用CDN服务隐藏真实服务器", "risk": "medium", "weight": 5},
        "REP_010": {"name": "DGA域名检测", "desc": "使用DGA生成的域名", "risk": "high", "weight": 7},
    },
    "开发者画像": {
        "DEV_001": {"name": "开发者认证缺失", "desc": "开发者未完成实名认证", "risk": "medium", "weight": 4},
        "DEV_002": {"name": "开发者资质存疑", "desc": "开发者资质信息异常", "risk": "high", "weight": 5},
        "DEV_003": {"name": "企业开发者异常", "desc": "企业开发者信息不一致", "risk": "medium", "weight": 4},
        "DEV_004": {"name": "开发者关联黑名单", "desc": "开发者关联到黑名单", "risk": "critical", "weight": 8},
        "DEV_005": {"name": "跨账号异常关联", "desc": "多个开发者账号存在异常关联", "risk": "high", "weight": 6},
        "DEV_006": {"name": "历史应用违规记录", "desc": "开发者有应用违规历史", "risk": "high", "weight": 6},
        "DEV_007": {"name": "频繁更换账号", "desc": "开发者频繁更换注册账号", "risk": "medium", "weight": 4},
        "DEV_008": {"name": "应用下架率高", "desc": "开发者应用下架率异常高", "risk": "high", "weight": 6},
        "DEV_009": {"name": "新开发者注册", "desc": "开发者账号注册时间过短", "risk": "low", "weight": 2},
        "DEV_010": {"name": "开发者活跃度异常", "desc": "开发者活跃度呈现异常模式", "risk": "medium", "weight": 3},
        "DEV_011": {"name": "批量发布相似应用", "desc": "批量发布功能相似的应用", "risk": "medium", "weight": 4},
        "DEV_012": {"name": "应用抄袭嫌疑", "desc": "应用代码与其他应用高度相似", "risk": "high", "weight": 6},
        "DEV_013": {"name": "热更新异常", "desc": "使用非官方渠道进行热更新", "risk": "high", "weight": 7},
        "DEV_014": {"name": "代码托管异常", "desc": "代码托管在非可信平台", "risk": "medium", "weight": 4},
        "DEV_015": {"name": "跨平台恶意分发", "desc": "同一应用在多平台恶意分发", "risk": "high", "weight": 6},
    }
}

# 数据源定义（横轴）
DATA_SOURCES = [
    {"id": "DS_01", "name": "安装包解析", "desc": "解析HAP包、module.json5、签名信息等"},
    {"id": "DS_02", "name": "静态代码分析", "desc": "反编译分析、代码扫描、模式匹配"},
    {"id": "DS_03", "name": "动态沙箱分析", "desc": "运行时行为监控、API调用追踪"},
    {"id": "DS_04", "name": "网络流量分析", "desc": "抓包分析、域名/IP检测、通信行为"},
    {"id": "DS_05", "name": "权限行为监控", "desc": "权限申请和使用监控"},
    {"id": "DS_06", "name": "数据流分析", "desc": "敏感数据追踪、存储分析"},
    {"id": "DS_07", "name": "用户评价分析", "desc": "应用商店评论、评分分析"},
    {"id": "DS_08", "name": "舆情监控", "desc": "社交媒体、新闻、公告监控"},
    {"id": "DS_09", "name": "威胁情报查询", "desc": " VirusTotal、恶意样本库、信誉库"},
    {"id": "DS_10", "name": "开发者信息", "desc": "开发者资质、历史记录、行为分析"},
    {"id": "DS_11", "name": "合规规则检查", "desc": "政策法规、行业规范匹配"},
]

# 标签与数据源的映射关系
# 格式: {标签ID: [数据源ID列表]}
LABEL_SOURCE_MAPPING = {
    # 应用基本信息画像 - 主要来自安装包解析和静态代码分析
    "APP_001": ["DS_01"], "APP_002": ["DS_01"], "APP_003": ["DS_01"], "APP_004": ["DS_01"], "APP_005": ["DS_01"],
    "APP_006": ["DS_01"], "APP_007": ["DS_01"], "APP_008": ["DS_01"], "APP_009": ["DS_01"], "APP_010": ["DS_01"],
    "APP_011": ["DS_01"], "APP_012": ["DS_01"], "APP_013": ["DS_01"], "APP_014": ["DS_01"], "APP_015": ["DS_01"],
    "CS_001": ["DS_02"], "CS_002": ["DS_02"], "CS_003": ["DS_01", "DS_02"], "CS_004": ["DS_01", "DS_02"],
    "CS_005": ["DS_02"], "CS_006": ["DS_02"], "CS_007": ["DS_02"], "CS_008": ["DS_02"], "CS_009": ["DS_02"],
    "CS_010": ["DS_02", "DS_03"], "CS_011": ["DS_02"], "CS_012": ["DS_02", "DS_09"], "CS_013": ["DS_02"],
    "CS_014": ["DS_01", "DS_02"], "CS_015": ["DS_02"],

    # 应用行为画像 - 主要来自动态分析
    "BH_001": ["DS_03"], "BH_002": ["DS_03"], "BH_003": ["DS_03"], "BH_004": ["DS_03"], "BH_005": ["DS_03"],
    "BH_006": ["DS_03", "DS_05"], "BH_007": ["DS_03", "DS_05"], "BH_008": ["DS_03", "DS_05"], "BH_009": ["DS_03", "DS_05"],
    "BH_010": ["DS_03"], "BH_011": ["DS_03", "DS_04"], "BH_012": ["DS_03"], "BH_013": ["DS_03"], "BH_014": ["DS_03"],
    "BH_015": ["DS_03"], "BH_016": ["DS_03"], "BH_017": ["DS_03"], "BH_018": ["DS_03"], "BH_019": ["DS_03"],
    "BH_020": ["DS_03"], "BH_021": ["DS_03"], "BH_022": ["DS_03"], "BH_023": ["DS_03"], "BH_024": ["DS_03"],
    "BH_025": ["DS_03", "DS_11"],

    # 网络通信画像
    "NET_001": ["DS_04"], "NET_002": ["DS_02", "DS_04"], "NET_003": ["DS_02", "DS_04"], "NET_004": ["DS_04"],
    "NET_005": ["DS_02", "DS_04"], "NET_006": ["DS_04", "DS_09"], "NET_007": ["DS_04"], "NET_008": ["DS_04"],
    "NET_009": ["DS_04", "DS_06"], "NET_010": ["DS_04"], "NET_011": ["DS_04", "DS_06"], "NET_012": ["DS_04", "DS_06"],
    "NET_013": ["DS_04"], "NET_014": ["DS_04", "DS_06"], "NET_015": ["DS_04"],

    # 数据安全画像
    "DS_001": ["DS_06"], "DS_002": ["DS_06"], "DS_003": ["DS_06"], "DS_004": ["DS_06"], "DS_005": ["DS_06"],
    "DS_006": ["DS_02"], "DS_007": ["DS_02"], "DS_008": ["DS_02"], "DS_009": ["DS_02"], "DS_010": ["DS_02", "DS_06"],
    "DS_011": ["DS_03"], "DS_012": ["DS_02"], "DS_013": ["DS_02"], "DS_014": ["DS_02"], "DS_015": ["DS_03"],

    # 权限使用画像
    "PERM_001": ["DS_01", "DS_05"], "PERM_002": ["DS_01", "DS_05"], "PERM_003": ["DS_01"], "PERM_004": ["DS_03"],
    "PERM_005": ["DS_01"], "PERM_006": ["DS_05"], "PERM_007": ["DS_05"], "PERM_008": ["DS_05"], "PERM_009": ["DS_05"],
    "PERM_010": ["DS_06"], "PERM_011": ["DS_05"], "PERM_012": ["DS_05"], "PERM_013": ["DS_03"], "PERM_014": ["DS_05"],
    "PERM_015": ["DS_03", "DS_05"],

    # 合规风险画像
    "COMP_001": ["DS_11"], "COMP_002": ["DS_11"], "COMP_003": ["DS_11"], "COMP_004": ["DS_11"], "COMP_005": ["DS_11"],
    "COMP_006": ["DS_01", "DS_11"], "COMP_007": ["DS_01", "DS_11"], "COMP_008": ["DS_11"], "COMP_009": ["DS_04", "DS_11"],
    "COMP_010": ["DS_11"],

    # 应用舆情画像
    "PB_001": ["DS_07"], "PB_002": ["DS_07"], "PB_003": ["DS_07"], "PB_004": ["DS_07"], "PB_005": ["DS_07"],
    "PB_006": ["DS_08"], "PB_007": ["DS_08", "DS_09"], "PB_008": ["DS_08"], "PB_009": ["DS_08"], "PB_010": ["DS_08"],
    "REP_001": ["DS_09"], "REP_002": ["DS_09"], "REP_003": ["DS_09"], "REP_004": ["DS_09"], "REP_005": ["DS_09"],
    "REP_006": ["DS_04", "DS_09"], "REP_007": ["DS_04", "DS_09"], "REP_008": ["DS_09"], "REP_009": ["DS_04"],
    "REP_010": ["DS_04"],

    # 开发者画像
    "DEV_001": ["DS_10"], "DEV_002": ["DS_10"], "DEV_003": ["DS_10"], "DEV_004": ["DS_10"], "DEV_005": ["DS_10"],
    "DEV_006": ["DS_10"], "DEV_007": ["DS_10"], "DEV_008": ["DS_10"], "DEV_009": ["DS_10"], "DEV_010": ["DS_10"],
    "DEV_011": ["DS_10"], "DEV_012": ["DS_02"], "DEV_013": ["DS_03"], "DEV_014": ["DS_10"], "DEV_015": ["DS_10"],
}


def create_styles():
    """创建单元格样式"""
    # 标题样式
    title_font = Font(name='微软雅黑', size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 表头样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 分类标题样式
    category_font = Font(name='微软雅黑', size=12, bold=True, color="FFFFFF")
    category_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    category_alignment = Alignment(horizontal='left', vertical='center')

    # 数据样式
    data_font = Font(name='微软雅黑', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    return {
        'title_font': title_font, 'title_fill': title_fill, 'title_alignment': title_alignment,
        'header_font': header_font, 'header_fill': header_fill, 'header_alignment': header_alignment,
        'category_font': category_font, 'category_fill': category_fill, 'category_alignment': category_alignment,
        'data_font': data_font, 'data_alignment': data_alignment,
        'thin_border': thin_border
    }


def create_labels_sheet(ws, styles):
    """创建标签列表Sheet"""
    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 20

    # 标题行
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "HarmonyOS 应用画像标签库 (145个核心标签)"
    title_cell.font = styles['title_font']
    title_cell.fill = styles['title_fill']
    title_cell.alignment = styles['title_alignment']
    title_cell.border = styles['thin_border']

    # 表头
    headers = ['标签ID', '标签名称', '描述', '风险等级', '权重', '所属维度']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_alignment']
        cell.border = styles['thin_border']

    # 填充数据
    current_row = 3
    for category, labels in LABELS_DATA.items():
        # 分类标题行
        ws.merge_cells(f'A{current_row}:F{current_row}')
        category_cell = ws[f'A{current_row}']
        category_cell.value = f"【{category}】"
        category_cell.font = styles['category_font']
        category_cell.fill = styles['category_fill']
        category_cell.alignment = styles['category_alignment']
        category_cell.border = styles['thin_border']
        current_row += 1

        # 标签数据行
        for label_id, label_info in labels.items():
            ws.cell(row=current_row, column=1, value=label_id)
            ws.cell(row=current_row, column=2, value=label_info['name'])
            ws.cell(row=current_row, column=3, value=label_info['desc'])
            ws.cell(row=current_row, column=4, value=label_info['risk'])
            ws.cell(row=current_row, column=5, value=label_info['weight'])
            ws.cell(row=current_row, column=6, value=category)

            # 设置数据行样式
            for col in range(1, 7):
                cell = ws.cell(row=current_row, column=col)
                cell.font = styles['data_font']
                cell.alignment = styles['data_alignment']
                cell.border = styles['thin_border']

                # 风险等级颜色标记
                if col == 4:
                    if label_info['risk'] == 'critical':
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        cell.font = Font(name='微软雅黑', size=10, bold=True, color="FFFFFF")
                    elif label_info['risk'] == 'high':
                        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                    elif label_info['risk'] == 'medium':
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    elif label_info['risk'] == 'low':
                        cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

            current_row += 1

    # 冻结首行
    ws.freeze_panes = 'A3'


def create_detection_matrix_sheet(ws, styles):
    """创建检测方案矩阵Sheet"""
    # 设置列宽
    ws.column_dimensions['A'].width = 20
    for i in range(1, len(DATA_SOURCES) + 1):
        ws.column_dimensions[get_column_letter(i + 1)].width = 12

    # 标题
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = "应用画像标签检测方案矩阵（数据源 × 标签）"
    title_cell.font = styles['title_font']
    title_cell.fill = styles['title_fill']
    title_cell.alignment = styles['title_alignment']
    title_cell.border = styles['thin_border']

    # 第一行：数据源表头
    ws.cell(row=2, column=1, value="标签ID / 数据源")
    ws['A2'].font = styles['header_font']
    ws['A2'].fill = styles['header_fill']
    ws['A2'].alignment = styles['header_alignment']
    ws['A2'].border = styles['thin_border']

    for col, source in enumerate(DATA_SOURCES, 2):
        cell = ws.cell(row=2, column=col)
        cell.value = source['id'] + "\n" + source['name']
        cell.font = Font(name='微软雅黑', size=9, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = styles['thin_border']

    # 第二行：数据源描述
    for col, source in enumerate(DATA_SOURCES, 2):
        cell = ws.cell(row=3, column=col)
        cell.value = source['desc']
        cell.font = Font(name='微软雅黑', size=8)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = styles['thin_border']

    ws.merge_cells('A2:A3')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 填充检测矩阵数据
    current_row = 4
    for category, labels in LABELS_DATA.items():
        # 分类标题行
        ws.merge_cells(f'A{current_row}:K{current_row}')
        category_cell = ws[f'A{current_row}']
        category_cell.value = f"【{category}】"
        category_cell.font = styles['category_font']
        category_cell.fill = styles['category_fill']
        category_cell.alignment = styles['category_alignment']
        category_cell.border = styles['thin_border']
        current_row += 1

        # 标签行
        for label_id, label_info in labels.items():
            # 第一列：标签ID和名称
            label_cell = ws.cell(row=current_row, column=1)
            label_cell.value = f"{label_id}\n{label_info['name']}"
            label_cell.font = Font(name='微软雅黑', size=9)
            label_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            label_cell.border = styles['thin_border']

            # 风险等级背景色
            if label_info['risk'] == 'critical':
                label_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif label_info['risk'] == 'high':
                label_cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            elif label_info['risk'] == 'medium':
                label_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

            # 数据源列
            sources = LABEL_SOURCE_MAPPING.get(label_id, [])
            for col, source in enumerate(DATA_SOURCES, 2):
                cell = ws.cell(row=current_row, column=col)
                cell.border = styles['thin_border']

                if source['id'] in sources:
                    # 检测能力标记
                    cell.value = "●"
                    cell.font = Font(name='微软雅黑', size=14, color="008000")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.value = ""
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            current_row += 1

    # 冻结前3行和第1列
    ws.freeze_panes = 'B4'


def create_summary_sheet(ws, styles):
    """创建统计摘要Sheet"""
    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

    # 标题
    ws.merge_cells('A1:C1')
    title_cell = ws['A1']
    title_cell.value = "标签体系统计摘要"
    title_cell.font = styles['title_font']
    title_cell.fill = styles['title_fill']
    title_cell.alignment = styles['title_alignment']
    title_cell.border = styles['thin_border']

    # 表头
    headers = ['统计项', '数量', '占比']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_alignment']
        cell.border = styles['thin_border']

    current_row = 3

    # 维度分布
    total_labels = sum(len(labels) for labels in LABELS_DATA.values())
    ws.cell(row=current_row, column=1, value="【维度分布】")
    ws[f'A{current_row}'].font = Font(name='微软雅黑', size=11, bold=True)
    ws.merge_cells(f'A{current_row}:C{current_row}')
    current_row += 1

    for category, labels in LABELS_DATA.items():
        count = len(labels)
        percent = f"{count / total_labels * 100:.1f}%"
        ws.cell(row=current_row, column=1, value=category)
        ws.cell(row=current_row, column=2, value=count)
        ws.cell(row=current_row, column=3, value=percent)
        for col in range(1, 4):
            ws.cell(row=current_row, column=col).border = styles['thin_border']
            ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1

    current_row += 1

    # 风险等级分布
    ws.cell(row=current_row, column=1, value="【风险等级分布】")
    ws[f'A{current_row}'].font = Font(name='微软雅黑', size=11, bold=True)
    ws.merge_cells(f'A{current_row}:C{current_row}')
    current_row += 1

    risk_counts = {'critical': 0, 'high': 0, 'medium': 0, 'low': 0}
    for labels in LABELS_DATA.values():
        for label_info in labels.values():
            risk_counts[label_info['risk']] += 1

    risk_names = {'critical': 'Critical (严重)', 'high': 'High (高)', 'medium': 'Medium (中)', 'low': 'Low (低)'}
    for risk, count in risk_counts.items():
        percent = f"{count / total_labels * 100:.1f}%"
        ws.cell(row=current_row, column=1, value=risk_names[risk])
        ws.cell(row=current_row, column=2, value=count)
        ws.cell(row=current_row, column=3, value=percent)
        for col in range(1, 4):
            cell = ws.cell(row=current_row, column=col)
            cell.border = styles['thin_border']
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1

    current_row += 1

    # 数据源覆盖统计
    ws.cell(row=current_row, column=1, value="【数据源覆盖统计】")
    ws[f'A{current_row}'].font = Font(name='微软雅黑', size=11, bold=True)
    ws.merge_cells(f'A{current_row}:C{current_row}')
    current_row += 1

    # 统计每个数据源能检测的标签数量
    source_coverage = {}
    for source in DATA_SOURCES:
        count = sum(1 for sources in LABEL_SOURCE_MAPPING.values() if source['id'] in sources)
        source_coverage[source['name']] = count

    # 按检测数量排序
    sorted_sources = sorted(source_coverage.items(), key=lambda x: x[1], reverse=True)
    for source_name, count in sorted_sources:
        percent = f"{count / total_labels * 100:.1f}%"
        ws.cell(row=current_row, column=1, value=source_name)
        ws.cell(row=current_row, column=2, value=count)
        ws.cell(row=current_row, column=3, value=percent)
        for col in range(1, 4):
            cell = ws.cell(row=current_row, column=col)
            cell.border = styles['thin_border']
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1


def main():
    """主函数"""
    # 创建工作簿
    wb = openpyxl.Workbook()

    # 删除默认Sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # 创建样式
    styles = create_styles()

    # 创建标签列表Sheet
    labels_sheet = wb.create_sheet("标签列表")
    create_labels_sheet(labels_sheet, styles)

    # 创建检测方案矩阵Sheet
    matrix_sheet = wb.create_sheet("检测方案矩阵")
    create_detection_matrix_sheet(matrix_sheet, styles)

    # 创建统计摘要Sheet
    summary_sheet = wb.create_sheet("统计摘要")
    create_summary_sheet(summary_sheet, styles)

    # 保存文件
    output_path = "D:/workspace/harmony-app-profiles/docs/harmonyos_app_profiling_labels.xlsx"
    wb.save(output_path)

    total_labels = sum(len(labels) for labels in LABELS_DATA.values())
    print(f"Done! Generated {total_labels} labels across {len(DATA_SOURCES)} data sources.")
    print(f"Output: {output_path}")
    print(f"Sheets: Label List, Detection Matrix, Summary")


if __name__ == "__main__":
    main()
